from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os

# Paths
INPUT_FILE = r"sample_data.xlsx"
OUTPUT_FILE = r"sample_data_updated.xlsx"

# Load workbook and sheet
wb = load_workbook(INPUT_FILE)
ws = wb.active

# -------------------------------
# FUNCTIONS
# -------------------------------

def add_row(name, amount):
    max_row = ws.max_row + 1
    ws.cell(row=max_row, column=1).value = name
    ws.cell(row=max_row, column=2).value = amount
    print(f"Added row: {name}, {amount}")

def update_row(row, name=None, amount=None):
    if name is not None:
        ws.cell(row=row, column=1).value = name
    if amount is not None:
        ws.cell(row=row, column=2).value = amount
    print(f"Updated row {row} with Name={name} Amount={amount}")

def delete_row(row):
    ws.delete_rows(row)
    print(f"Deleted row {row}")

def format_header():
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['A1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws['B1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    print("Formatted header")

# -------------------------------
# MAIN AUTOMATION
# -------------------------------

format_header()
update_row(2, name="John Doe", amount=500)
delete_row(4)
add_row("Michael", 600)
add_row("Sophia", 700)
add_row("David", 800)

# Save as new file
wb.save(OUTPUT_FILE)
print(f"✅ Excel automation complete! Saved as: {OUTPUT_FILE}")
